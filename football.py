from time import sleep
import os
try:
    import openpyxl
    import requests
    import bs4
    from matplotlib import pyplot as plt
    import lxml
except ModuleNotFoundError:
    print('''You must install modules: requests, bs4, openpyxl, matplotlib, lxml.
If You don\'t know how to do this, look on this website:
https://docs.python.org/3/installing/index.html''')
    input()
    quit()


class Footballer:

    def __init__(self, name, club, club_country, cost, nationality):
        self.name = name
        self.club = club
        self.club_country = club_country
        self.cost = cost
        self.nationality = nationality

    def __str__(self):

        if '- loan' in self.club:
            return self.name + 'was loaned by ' + self.club[:-6] + 'for ' + str(self.cost) + ' mln €.'

        else:
            return self.name + 'was bought by ' + self.club + 'for ' + str(self.cost) + ' mln €.'


class ListOfTransfers:

    def __init__(self, year):

        self.year = year
        scraped_list = scraping_transfermarkt(self.year)
        self.footballers = []

        for key, values in scraped_list.items():
            player = Footballer(key, values[0][0], values[0][1], values[1], values[2])
            self.footballers.append(player)

    def sum_cost(self):

        all_value = 0

        for player in self.footballers:
            all_value += player.cost

        return all_value

    def sum_club_countries(self):

        all_club_countries = []

        for player in self.footballers:
            all_club_countries.append(player.club_country)

        club_countries = {country: all_club_countries.count(country) for country in all_club_countries}

        return club_countries

    def sum_nationalities(self):

        all_nationalities = []

        for player in self.footballers:
            all_nationalities.append(player.nationality)

        nationalities = {country: all_nationalities.count(country) for country in all_nationalities}

        return nationalities

    def no_loan(self):

        loans = 0

        for player in self.footballers:

            if '- loan' in player.club:
                loans += 1

        return loans


def scraping_transfermarkt(year):
    url = 'https://www.transfermarkt.com/transfers/saisontransfers/statistik/top/plus/0/galerie/0?saison_id={}&transferf' \
          'enster=alle&land_id=&ausrichtung=&spielerposition_id=&altersklasse=&leihe='
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:63.0) Gecko/20100101 Firefox/63.0'}

    result = requests.get(url.format(year), headers=headers)
    soup = bs4.BeautifulSoup(result.text, 'lxml')

    '''
    On the transfermarket page, footballers are divided in lines and marked with numbers
    (the first on the page - 1,the second - 2, the third - 3, ...), footballers with even numbers
    are assigned to a different class (in the source code of the html page) than those with odd numbers
    '''

    odd_class = soup.select('.odd')  # extracting data about a footballer (class = odd)
    even_class = soup.select('.even')  # extracting data about a footballer (class = even)

    list_of_footballer = {}
    odd = True

    for i in range(10):
        '''
        loop pulls out the data of the first 20 players on the page
        name - the name of the footballer,
        club - the club which bought/loaned the player
        cost - buyout amount
        club_country - the club's country of origin
        nationality - nationality of the player (in the case of two, only one is given, the first given on the page)
        '''

        if odd:
            name = odd_class[i // 2].select('.hauptlink')[0].text
            name = name[1:]
            club = odd_class[i // 2].select('.hauptlink')[1].text
            club = club[1:]
            cost = odd_class[i // 2].select('.hauptlink')[2].text

            club_country = odd_class[i // 2].select('.zentriert')[3].select('.flaggenrahmen')[0]['title']
            nationality = odd_class[i // 2].select('.zentriert')[2].select('.flaggenrahmen')[0]['title']

            odd = False

        elif not odd:
            name = even_class[i // 2].select('.hauptlink')[0].text
            name = name[1:]
            club = even_class[i // 2].select('.hauptlink')[1].text
            club = club[1:]
            cost = even_class[i // 2].select('.hauptlink')[2].text

            club_country = even_class[i // 2].select('.zentriert')[3].select('.flaggenrahmen')[0]['title']
            nationality = even_class[i // 2].select('.zentriert')[2].select('.flaggenrahmen')[0]['title']

            odd = True

        if 'Loan fee' in cost:
            cost = cost[10:]
            cost = float(cost[:-1])
            club = club + '- loan'

        elif 'transfer' in cost.lower():
            cost = 0.0

        else:
            cost = cost[1:]
            cost = float(cost[:-1])

        list_of_footballer[name] = [(club, club_country), cost, nationality]

    return list_of_footballer


def create_sheet(wb, players_class):

    sheet = wb.create_sheet(title=f'{players_class.year}')
    sheet['A1'] = 'Footballer'
    sheet['B1'] = 'Nation'
    sheet['C1'] = 'Club'
    sheet['D1'] = 'Cost [mln €]'

    i = 2
    for player in players_class.footballers:
        sheet[f'A{i}'] = player.name
        sheet[f'B{i}'] = player.nationality
        sheet[f'C{i}'] = player.club
        sheet[f'D{i}'] = player.cost
        i += 1

        if player == players_class.footballers[-1]:
            sheet[f'C{13}'] = 'SUMA'
            sheet[f'D{13}'] = players_class.sum_cost()
    return sheet


# data -> dictionary eg. {2010: 850,4} -> the amount of money spent in a given year on transfers (top 10)
def create_graph(data1, data2, title1, title2):

    fig, axis = plt.subplots(2, figsize=(7.5, 7.5), num='Best transfers')

    # first graph
    x1 = data1.keys()

    y1 = data1.values()
    graph1 = axis[0].bar(x1, y1, width=0.3)
    axis[0].set_xlabel('Season in which transfers were made')
    axis[0].set_ylabel('Total prices of players [€ million].')
    axis[0].bar_label(graph1)
    axis[0].set_title(title1)
    axis[0].set_ylim([0, max(data1.values())+150])

    # second graph
    x2 = data2.keys()
    y2 = data2.values()
    graph2 = axis[1].bar(x2, y2, width=0.6)
    axis[1].set_xlabel('Countries where footballers were bought.')
    axis[1].set_ylabel('Number of players bought.')
    axis[1].bar_label(graph2)
    axis[1].set_title(title2)
    axis[1].set_ylim([0, max(data2.values())+10])

    plt.subplots_adjust(hspace=1)

    plt.show()


# name -> string, years -> tuple: (initial year, final year)
def start_program(name, years):

    wb = openpyxl.Workbook()

    # data -> dictionary np. {2010: 850,4} -> the amount of money spent in a given year on transfers (top 10)
    sum_of_cost = {}

    # data -> dictionary np. {'England': 10} -> the number of transfers in clubs from a given country
    sum_club_countries = {}

    for year in range(years[0], years[1]+1):
        best_transfers = ListOfTransfers(year)
        create_sheet(wb, best_transfers)

        # data for graphs -> sum of players cost
        sum_of_cost[str(year)] = best_transfers.sum_cost()

        # data for graphs -> sum of transfers in specific country
        for key, item in best_transfers.sum_club_countries().items():

            if key in sum_club_countries:
                sum_club_countries[key] += item

            else:
                sum_club_countries[key] = item

    del wb['Sheet']
    wb.save(f'{name}.xlsx')
    wb.close()

    file = os.getcwd()
    os.startfile(f'{file}\\{name}.xlsx')
    sleep(2)

    create_graph(sum_of_cost, sum_club_countries, 'Sum of transfer prices in a given year',
                 'Number of transfers in a country')


if __name__=='__main__':
    pass